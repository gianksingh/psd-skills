#!/usr/bin/env python3
"""
PSD DC Labor Forecast workbook builder.

Reads a JSON config (see references/config_schema.md) and writes a .xlsx that
matches the PSD forecast template exactly:
  - one daily forecast tab per month (WEB = 17 cols w/ Y/Y; TTS = 8 cols simplified)
  - a "<period> Total" pacing dashboard (Forecast / LY / Y/Y / Actual / Pacing)
  - an Appendix tab (monthly controls, DOW weights[web], promo table[web], methodology)

The script owns ALL layout, formulas and styling so output is uniform.
Claude only supplies the numbers (daily net sales distribution, per-day AOV/UPT,
notes, LY comp denominators, monthly LY actuals, assumptions) in the config.

Usage:  python3 build_workbook.py config.json /path/to/output.xlsx
"""
import json
import sys
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- styling ----
NAVY = "1F2A44"          # header band (white bold text)
YELLOW_INPUT = "FFF2CC"  # forecast input cells (editable: net sales, AOV, UPT)
YELLOW_ACTUAL = "FFFF00"  # "Actual (fill in)" cells on the dashboard
GREY_TOTAL = "D9D9D9"     # total rows
GREY_CALC = "EFEFEF"      # calculated (non-editable) cells on dashboard

F_COUNT = "#,##0"
F_MONEY = "\\$#,##0"
F_AOV = "\\$0.00"
F_UPT = "0.00"
F_YOY = "\\+0%;\\-0%"
F_PCT1 = "0.0%"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def hdr_font():
    return Font(bold=True, color="FFFFFF", size=9)


def fill(hexcolor):
    # ensure opaque ARGB so Excel renders the fill (6-hex defaults to alpha 00 = transparent)
    if len(hexcolor) == 6:
        hexcolor = "FF" + hexcolor
    return PatternFill("solid", fgColor=hexcolor)


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.font = hdr_font()
        cell.fill = fill(NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def numstr(v):
    """Render a number for embedding in a formula: drop trailing .0 on whole numbers."""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def parse_d(s):
    if isinstance(s, (datetime, date)):
        return datetime(s.year, s.month, s.day)
    return datetime.strptime(s, "%Y-%m-%d")


# ------------------------------------------------------------ daily tabs ----
def build_web_daily(ws, m, appendix_col):
    """17-column WEB daily tab. appendix_col is the letter (B/C/D...) for this month."""
    title = m["daily_title"]
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = m["daily_subtitle"]
    ws["A2"].font = Font(italic=True, size=9, color="595959")

    ws.merge_cells("L3:N3")
    ws.merge_cells("O3:Q3")
    ws["L3"] = "vs SAME CALENDAR DATE LY"
    ws["O3"] = "vs SAME WEEK # + DAY LY (retail)"
    for ref in ("L3", "O3"):
        ws[ref].font = Font(bold=True, size=9)
        ws[ref].alignment = Alignment(horizontal="center")

    headers = ["Day", "Week + Day", "Orders", "Adj Net Sales", "Adj Units",
               "OP Units", "Adj AOV", "Adj UPT", "Ground Orders", "Exped. Orders",
               "Notes / DC Callouts", "Orders Y/Y", "Net Sales Y/Y", "Units Y/Y",
               "Orders Y/Y", "Net Sales Y/Y", "Units Y/Y"]
    for i, h in enumerate(headers, 1):
        ws.cell(4, i, h)
    style_header_row(ws, 4, 17)

    days = m["days"]
    r = 5
    for d in days:
        ws.cell(r, 1, parse_d(d["date"])).number_format = "m/d/yyyy"
        ws.cell(r, 2, f'="Wk-"&WEEKNUM(A{r},1)&" "&TEXT(A{r},"ddd")')
        ws.cell(r, 3, f"=ROUND(D{r}/G{r},0)").number_format = F_COUNT
        c_ns = ws.cell(r, 4, d["net_sales"]); c_ns.number_format = F_MONEY
        ws.cell(r, 5, f"=ROUND(C{r}*H{r},0)").number_format = F_COUNT
        ws.cell(r, 6, f"=ROUND(C{r}*Appendix!{appendix_col}$8,0)").number_format = F_COUNT
        c_aov = ws.cell(r, 7, d["aov"]); c_aov.number_format = F_AOV
        c_upt = ws.cell(r, 8, d["upt"]); c_upt.number_format = F_UPT
        ws.cell(r, 9, f"=C{r}-J{r}").number_format = F_COUNT
        ws.cell(r, 10, f"=ROUND(C{r}*Appendix!{appendix_col}$9,0)").number_format = F_COUNT
        if d.get("note"):
            ws.cell(r, 11, d["note"])
        # editable forecast inputs -> yellow
        for cell in (c_ns, c_aov, c_upt):
            cell.fill = fill(YELLOW_INPUT)
        cal = d["cal_ly"]; ret = d["retail_ly"]
        ws.cell(r, 12, f"=C{r}/{numstr(cal['orders'])}-1").number_format = F_YOY
        ws.cell(r, 13, f"=D{r}/{numstr(cal['net_sales'])}-1").number_format = F_YOY
        ws.cell(r, 14, f"=E{r}/{numstr(cal['units'])}-1").number_format = F_YOY
        ws.cell(r, 15, f"=C{r}/{numstr(ret['orders'])}-1").number_format = F_YOY
        ws.cell(r, 16, f"=D{r}/{numstr(ret['net_sales'])}-1").number_format = F_YOY
        ws.cell(r, 17, f"=E{r}/{numstr(ret['units'])}-1").number_format = F_YOY
        r += 1

    last = r - 1
    tr = r  # total row
    ws.cell(tr, 1, m["total_label"])
    ws.cell(tr, 3, f"=SUM(C5:C{last})").number_format = F_COUNT
    ws.cell(tr, 4, f"=SUM(D5:D{last})").number_format = F_MONEY
    ws.cell(tr, 5, f"=SUM(E5:E{last})").number_format = F_COUNT
    ws.cell(tr, 6, f"=SUM(F5:F{last})").number_format = F_COUNT
    ws.cell(tr, 7, f"=D{tr}/C{tr}").number_format = F_AOV
    ws.cell(tr, 8, f"=E{tr}/C{tr}").number_format = F_UPT
    ws.cell(tr, 9, f"=SUM(I5:I{last})").number_format = F_COUNT
    ws.cell(tr, 10, f"=SUM(J5:J{last})").number_format = F_COUNT
    mt = m["month_ly"]
    ws.cell(tr, 12, f"=C{tr}/{numstr(mt['orders'])}-1").number_format = F_YOY
    ws.cell(tr, 13, f"=D{tr}/{numstr(mt['net_sales'])}-1").number_format = F_YOY
    ws.cell(tr, 14, f"=E{tr}/{numstr(mt['units'])}-1").number_format = F_YOY
    ws.cell(tr, 15, f"=C{tr}/{numstr(mt['retail_orders'])}-1").number_format = F_YOY
    ws.cell(tr, 16, f"=D{tr}/{numstr(mt['retail_net_sales'])}-1").number_format = F_YOY
    ws.cell(tr, 17, f"=E{tr}/{numstr(mt['retail_units'])}-1").number_format = F_YOY
    for c in range(1, 18):
        tc = ws.cell(tr, c)
        tc.font = Font(bold=True, size=10)
        tc.fill = fill(GREY_TOTAL)

    widths = {"A": 11, "B": 10, "C": 9, "D": 12, "E": 9, "F": 9, "G": 8, "H": 8,
              "I": 9, "J": 9, "K": 46, "L": 9, "M": 11, "N": 9, "O": 9, "P": 11, "Q": 9}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "C5"
    return tr


def build_tts_daily(ws, m, appendix_col):
    """8-column simplified TTS daily tab (no Y/Y, no OP/ship split)."""
    ws["A1"] = m["daily_title"]
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = m["daily_subtitle"]
    ws["A2"].font = Font(italic=True, size=9, color="595959")

    headers = ["Day", "Week + Day", "Orders", "Adj Net Sales", "Adj Units",
               "Adj AOV", "Adj UPT", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(4, i, h)
    style_header_row(ws, 4, 8)

    days = m["days"]
    r = 5
    for d in days:
        ws.cell(r, 1, parse_d(d["date"])).number_format = "m/d/yyyy"
        ws.cell(r, 2, f'="Wk-"&WEEKNUM(A{r},1)&" "&TEXT(A{r},"ddd")')
        ws.cell(r, 3, f"=ROUND(D{r}/F{r},0)").number_format = F_COUNT
        c_ns = ws.cell(r, 4, d["net_sales"]); c_ns.number_format = F_MONEY
        c_ns.fill = fill(YELLOW_INPUT)
        ws.cell(r, 5, f"=ROUND(C{r}*G{r},0)").number_format = F_COUNT
        ws.cell(r, 6, f"=Appendix!{appendix_col}$6").number_format = F_AOV
        ws.cell(r, 7, f"=Appendix!{appendix_col}$7").number_format = F_UPT
        if d.get("note"):
            ws.cell(r, 8, d["note"])
        r += 1

    last = r - 1
    tr = r
    ws.cell(tr, 1, m["total_label"])
    ws.cell(tr, 3, f"=SUM(C5:C{last})").number_format = F_COUNT
    ws.cell(tr, 4, f"=SUM(D5:D{last})").number_format = F_MONEY
    ws.cell(tr, 5, f"=SUM(E5:E{last})").number_format = F_COUNT
    ws.cell(tr, 6, f"=D{tr}/C{tr}").number_format = F_AOV
    ws.cell(tr, 7, f"=E{tr}/C{tr}").number_format = F_UPT
    for c in range(1, 9):
        tc = ws.cell(tr, c)
        tc.font = Font(bold=True, size=10)
        tc.fill = fill(GREY_TOTAL)

    widths = {"A": 11, "B": 10, "C": 9, "D": 12, "E": 9, "F": 8, "G": 8, "H": 40}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "C5"
    return tr


# ----------------------------------------------------- Q3 Total dashboard ----
# metric -> daily-tab total-row column letter
WEB_METRICS = [
    ("Adj Net Sales", "D", F_MONEY, "sum"),
    ("Orders", "C", F_COUNT, "sum"),
    ("Adj Units", "E", F_COUNT, "sum"),
    ("Adj AOV", "G", F_AOV, "aov"),
    ("Adj UPT", "H", F_UPT, "upt"),
    ("OP Units", "F", F_COUNT, "sum"),
    ("Ground Orders", "I", F_COUNT, "sum"),
    ("Expedited Orders", "J", F_COUNT, "sum"),
]
TTS_METRICS = [
    ("Adj Net Sales", "D", F_MONEY, "sum"),
    ("Orders", "C", F_COUNT, "sum"),
    ("Adj Units", "E", F_COUNT, "sum"),
    ("Adj AOV", "F", F_AOV, "aov"),
    ("Adj UPT", "G", F_UPT, "upt"),
]

DASH_HEADERS = ["Metric", "Forecast", None, "Fcst Y/Y %", "Actual (fill in)",
                "Pacing % to Fcst", "Actual Y/Y %"]


def build_dashboard(ws, cfg, month_total_rows):
    channel = cfg["channel"]
    metrics = WEB_METRICS if channel == "web" else TTS_METRICS
    nm = len(metrics)
    ly_label = "LY Actual (2025)" if channel == "web" else "LY Actual (2025, ref only)"

    ws["A1"] = cfg["summary_title"]
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = cfg["summary_subtitle"]
    ws["A2"].font = Font(italic=True, size=9, color="595959")

    months = cfg["months"]
    period = cfg["period_label"]
    block = 2 + nm                # header row + metric rows
    # total block at row 4; month blocks follow with 1 blank-row gap
    total_start = 4
    month_starts = []
    s = total_start + block + 1
    for _ in months:
        month_starts.append(s)
        s += block + 1
    # map metric index -> row within each month block
    def metric_row(block_start, idx):
        return block_start + 2 + idx  # +1 title, +1 header

    def write_header(row, title):
        ws.cell(row, 1, title).font = Font(bold=True, size=11)
        hr = row + 1
        labels = ["Metric", "Forecast", ly_label, "Fcst Y/Y %",
                  "Actual (fill in)", "Pacing % to Fcst", "Actual Y/Y %"]
        for i, lab in enumerate(labels, 1):
            ws.cell(hr, i, lab)
        style_header_row(ws, hr, 7)

    # ---- month blocks ----
    for mi, (m, bstart) in enumerate(zip(months, month_starts)):
        write_header(bstart, m["dashboard_title"])
        sheet = m["sheet_name"]
        trow = month_total_rows[mi]
        for idx, (name, col, fmt, _) in enumerate(metrics):
            r = metric_row(bstart, idx)
            ws.cell(r, 1, name)
            ws.cell(r, 2, f"='{sheet}'!{col}{trow}").number_format = fmt
            ly = m.get("ly_actual", {})
            ly_key = {"Adj Net Sales": "net_sales", "Orders": "orders",
                      "Adj Units": "units", "Adj AOV": "aov", "Adj UPT": "upt"}.get(name)
            if ly_key and ly.get(ly_key) is not None:
                ws.cell(r, 3, ly[ly_key]).number_format = fmt
            ws.cell(r, 4, f'=IF(C{r}="","",B{r}/C{r}-1)').number_format = F_YOY
            ec = ws.cell(r, 5); ec.fill = fill(YELLOW_ACTUAL); ec.number_format = fmt
            ws.cell(r, 6, f'=IF(E{r}="","",E{r}/B{r})').number_format = F_PCT1
            ws.cell(r, 7, f'=IF(OR(E{r}="",C{r}=""),"",E{r}/C{r}-1)').number_format = F_YOY

    # ---- total block (references month blocks) ----
    write_header(total_start, f"{period} TOTAL")
    # build, per metric, the list of month-block cell refs
    def refs(idx, letter):
        return [f"{letter}{metric_row(bs, idx)}" for bs in month_starts]
    # need orders index for aov/upt ratios
    orders_idx = next(i for i, mm in enumerate(metrics) if mm[0] == "Orders")
    ns_idx = next(i for i, mm in enumerate(metrics) if mm[0] == "Adj Net Sales")
    units_idx = next(i for i, mm in enumerate(metrics) if mm[0] == "Adj Units")

    for idx, (name, col, fmt, kind) in enumerate(metrics):
        r = total_start + 2 + idx
        ws.cell(r, 1, name)
        if kind == "sum":
            ws.cell(r, 2, "=" + "+".join(refs(idx, "B"))).number_format = fmt
            # LY total only where months carry LY for this metric
            if name in ("Adj Net Sales", "Orders", "Adj Units"):
                ws.cell(r, 3, "=" + "+".join(refs(idx, "C"))).number_format = fmt
            e_sum = "+".join(refs(idx, "E"))
            e_blank = ",".join([f'{c}=""' for c in refs(idx, "E")])
            ws.cell(r, 5, f'=IF(OR({e_blank}),"",{e_sum})').number_format = fmt
        elif kind == "aov":
            ws.cell(r, 2, f"=({'+'.join(refs(ns_idx,'B'))})/({'+'.join(refs(orders_idx,'B'))})").number_format = fmt
            ws.cell(r, 3, f"=({'+'.join(refs(ns_idx,'C'))})/({'+'.join(refs(orders_idx,'C'))})").number_format = fmt
            e_ns = "+".join(refs(ns_idx, "E")); e_or = "+".join(refs(orders_idx, "E"))
            e_blank = ",".join([f'{c}=""' for c in refs(ns_idx, "E")])
            ws.cell(r, 5, f'=IF(OR({e_blank}),"",({e_ns})/({e_or}))').number_format = fmt
        elif kind == "upt":
            ws.cell(r, 2, f"=({'+'.join(refs(units_idx,'B'))})/({'+'.join(refs(orders_idx,'B'))})").number_format = fmt
            ws.cell(r, 3, f"=({'+'.join(refs(units_idx,'C'))})/({'+'.join(refs(orders_idx,'C'))})").number_format = fmt
            e_un = "+".join(refs(units_idx, "E")); e_or = "+".join(refs(orders_idx, "E"))
            e_blank = ",".join([f'{c}=""' for c in refs(units_idx, "E")])
            ws.cell(r, 5, f'=IF(OR({e_blank}),"",({e_un})/({e_or}))').number_format = fmt
        ws.cell(r, 4, f'=IF(C{r}="","",B{r}/C{r}-1)').number_format = F_YOY
        ws.cell(r, 5).fill = fill(GREY_CALC)
        ws.cell(r, 6, f'=IF(E{r}="","",E{r}/B{r})').number_format = F_PCT1
        ws.cell(r, 7, f'=IF(OR(E{r}="",C{r}=""),"",E{r}/C{r}-1)').number_format = F_YOY

    widths = {"A": 18, "B": 14, "C": 18, "D": 12, "E": 14, "F": 15, "G": 13}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ------------------------------------------------------------- Appendix ----
def build_appendix(ws, cfg):
    channel = cfg["channel"]
    ap = cfg["appendix"]
    months = cfg["months"]
    ws["A1"] = ap["title"]
    ws["A1"].font = Font(bold=True, size=13)

    ws["A3"] = "MONTHLY CONTROLS"
    ws["A3"].font = Font(bold=True)
    for i, m in enumerate(months):
        ws.cell(4, 2 + i, m["name"]).font = Font(bold=True)
    rows = [
        ("Net Sales Target ($, adjusted)", "net_sales_target", F_MONEY, ap.get("note_target")),
        ("Baseline AOV ($, adjusted)" if channel == "web" else "Adj AOV ($, constant)",
         "baseline_aov", F_AOV, ap.get("note_aov")),
        ("Baseline UPT (adjusted)" if channel == "web" else "Adj UPT (constant)",
         "baseline_upt", F_UPT, ap.get("note_upt")),
    ]
    if channel == "web":
        rows += [
            ("OP Attach Rate", "op_attach", F_PCT1, ap.get("note_op")),
            ("Expedited % of Orders", "expedited_pct", F_PCT1, ap.get("note_expedited")),
        ]
    r = 5
    for label, key, fmt, note in rows:
        ws.cell(r, 1, label)
        for i, m in enumerate(months):
            ws.cell(r, 2 + i, m[key]).number_format = fmt
        if note:
            ws.cell(r, 5, note).font = Font(italic=True, size=9, color="595959")
        r += 1

    if channel == "web":
        r = 12
        ws.cell(r, 1, ap.get("dow_header", "DAY-OF-WEEK BASELINE WEIGHTS (share of week)")).font = Font(bold=True)
        r += 1
        for day in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]:
            ws.cell(r, 1, day)
            ws.cell(r, 2, ap["dow_weights"][day]).number_format = "0.0000"
            r += 1
        r += 2  # two blank rows before promo block (matches template)
        ws.cell(r, 1, ap.get("promo_header", "PROMO EVENT ASSUMPTIONS")).font = Font(bold=True)
        r += 1
        ph = ["Event", "2026 Dates", "Daily Lift vs Baseline", "Promo AOV", "Promo UPT / Notes"]
        for i, h in enumerate(ph, 1):
            ws.cell(r, i, h).font = Font(bold=True, size=9)
        r += 1
        for ev in ap.get("promo_events", []):
            ws.cell(r, 1, ev["event"])
            ws.cell(r, 2, ev["dates"])
            ws.cell(r, 3, ev["lift"])
            ws.cell(r, 4, ev["aov"])
            ws.cell(r, 5, ev["notes"])
            r += 1
        r += 2  # two blank rows before METHODOLOGY (matches template)
    else:
        r = 10

    ws.cell(r, 1, "METHODOLOGY").font = Font(bold=True)
    r += 1
    for line in ap["methodology"]:
        ws.cell(r, 1, line)
        r += 1

    ws.column_dimensions["A"].width = 42
    for col in ("B", "C", "D"):
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["E"].width = 60


# ----------------------------------------------------------------- main ----
def main():
    cfg_path, out_path = sys.argv[1], sys.argv[2]
    with open(cfg_path) as f:
        cfg = json.load(f)
    channel = cfg["channel"]
    assert channel in ("web", "tts"), "channel must be 'web' or 'tts'"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Dashboard first (so it's the leftmost tab, matching template)
    dash = wb.create_sheet(cfg["dashboard_sheet_name"])

    month_total_rows = []
    for i, m in enumerate(cfg["months"]):
        ws = wb.create_sheet(m["sheet_name"])
        ap_col = get_column_letter(2 + i)  # B, C, D ...
        if channel == "web":
            tr = build_web_daily(ws, m, ap_col)
        else:
            tr = build_tts_daily(ws, m, ap_col)
        month_total_rows.append(tr)

    ap_ws = wb.create_sheet("Appendix")
    build_appendix(ap_ws, cfg)

    build_dashboard(dash, cfg, month_total_rows)

    wb.save(out_path)
    print(f"Wrote {out_path} ({channel}, {len(cfg['months'])} month(s))")


if __name__ == "__main__":
    main()
